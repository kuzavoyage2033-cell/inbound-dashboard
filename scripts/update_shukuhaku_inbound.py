#!/usr/bin/env python3
"""
宿泊旅行統計 参考第１表（インバウンド）月次自動更新スクリプト
- MLITサイトから第２次速報の最新集計結果をダウンロード
- 対象ファイルの同月データを上書き（既存データがあれば削除して追加）
- 15箇所ランダム検証

使い方:
  python3 update_shukuhaku_inbound.py
  python3 update_shukuhaku_inbound.py --dry-run   # 取得のみ、ファイル更新しない
"""

import sys
import re
import random
import io
import argparse
from datetime import datetime
from html.parser import HTMLParser
from pathlib import Path
from urllib.request import urlopen, Request
from urllib.error import URLError

import pandas as pd
from openpyxl import load_workbook

MLIT_URL = "https://www.mlit.go.jp/kankocho/tokei_hakusyo/shukuhakutokei.html"
BASE_URL = "https://www.mlit.go.jp"
_BASE_DIR = Path(__file__).resolve().parent.parent
DATA_DIR = _BASE_DIR.parent.parent / "00_観光データ格納庫"
DL_FOLDER = str(DATA_DIR / '宿泊データ自動DL')

def _find_tate_file():
    """宿泊データ自動DL内の縦持ちファイルを返す（最新日付のもの）"""
    import re, unicodedata
    dl_dir = DATA_DIR / '宿泊データ自動DL'
    files = sorted(
        p for p in dl_dir.iterdir()
        if re.match(r'\d{8}_', p.name)
        and p.name.endswith('.xlsx')
        and '縦持ち' in unicodedata.normalize('NFC', p.name)
    )
    if not files:
        raise FileNotFoundError("縦持ちファイルが見つかりません: " + str(dl_dir))
    return str(files[-1])

EXCEL_FILE = _find_tate_file()

NATIONALITY_MAP = {
    "韓国": "韓国", "中国": "中国", "香港": "香港", "台湾": "台湾",
    "米国": "アメリカ", "カナダ": "カナダ", "英国": "イギリス",
    "ドイツ": "ドイツ", "フランス": "フランス", "ロシア": "ロシア",
    "シンガポール": "シンガポール", "タイ": "タイ", "マレーシア": "マレーシア",
    "インド": "インド", "オーストラリア": "オーストラリア", "インドネシア": "インドネシア",
    "ベトナム": "ベトナム", "フィリピン": "フィリピン", "イタリア": "イタリア",
    "スペイン": "スペイン", "北欧地域": "北欧地域", "中東地域": "中東地域",
    "メキシコ": "メキシコ", "その他": "その他",
}


class LinkParser(HTMLParser):
    """第２次速報セクション内の「集計結果」xlsxリンクを収集"""
    def __init__(self):
        super().__init__()
        self.in_section = False
        self.links = []       # [(label, href), ...]
        self._current_href = None
        self._current_text = ""
        self._depth = 0
        self._section_keyword_found = False

    def handle_starttag(self, tag, attrs):
        attrs_dict = dict(attrs)
        if tag == "a" and "href" in attrs_dict:
            self._current_href = attrs_dict["href"]
            self._current_text = ""

    def handle_data(self, data):
        if self._current_href is not None:
            self._current_text += data
        # 第２次速報セクションに入ったか判定
        if "第２次速報" in data or "第2次速報" in data:
            self._section_keyword_found = True

    def handle_endtag(self, tag):
        if tag == "a" and self._current_href is not None:
            text = self._current_text.strip()
            href = self._current_href
            if href.endswith(".xlsx") and "集計結果" in text:
                self.links.append((text, href))
            self._current_href = None
            self._current_text = ""


def fetch_latest_excel_url():
    """MLITページから第２次速報の最新集計結果xlsxのURLを返す"""
    print(f"[1] MLITページを取得中: {MLIT_URL}")
    req = Request(MLIT_URL, headers={"User-Agent": "Mozilla/5.0"})
    with urlopen(req, timeout=30) as resp:
        html = resp.read().decode("utf-8", errors="replace")

    # 第２次速報セクション以降のHTMLを抽出
    idx = html.find("第２次速報")
    if idx == -1:
        idx = html.find("第2次速報")
    section_html = html[idx:] if idx != -1 else html

    # 集計結果 xlsx リンクを全取得
    parser = LinkParser()
    parser.feed(section_html)

    if not parser.links:
        # フォールバック: 正規表現で /kankocho/content/*.xlsx を直接抽出
        xlsx_links = re.findall(r'/kankocho/content/(\d+)\.xlsx', section_html)
        if not xlsx_links:
            raise RuntimeError("集計結果リンクが見つかりませんでした")
        # 番号が最大のもの（= 最新）
        latest_id = max(xlsx_links, key=lambda x: int(x))
        url = f"{BASE_URL}/kankocho/content/{latest_id}.xlsx"
        print(f"   (正規表現フォールバック) 最新ファイル: {latest_id}.xlsx")
        return url, f"{latest_id}.xlsx"

    # 番号が最大のリンクを最新とみなす
    def extract_id(href):
        m = re.search(r'/(\d+)\.xlsx$', href)
        return int(m.group(1)) if m else 0

    latest_label, latest_href = max(parser.links, key=lambda x: extract_id(x[1]))
    url = BASE_URL + latest_href
    filename = latest_href.split("/")[-1]
    print(f"   最新集計結果: {latest_label} → {filename}")
    return url, filename


def download_excel(url):
    """URLからExcelをバイト列でダウンロード"""
    print(f"[2] ダウンロード中: {url}")
    req = Request(url, headers={"User-Agent": "Mozilla/5.0"})
    with urlopen(req, timeout=60) as resp:
        data = resp.read()
    print(f"   サイズ: {len(data)/1024:.1f} KB")
    return data


def find_ref1_sheet(xl_sheets):
    """参考第1表(X月) のシート名を返す（最初に見つかったもの）"""
    for name in xl_sheets:
        if re.match(r'参考第1表\(\d+月\)', name):
            return name
    return None


def parse_month_from_sheet(sheet_name):
    """参考第1表(2月) → 2"""
    m = re.search(r'\((\d+)月\)', sheet_name)
    return int(m.group(1)) if m else None


def parse_value(v):
    if pd.isna(v):
        return None
    s = str(v).strip().replace(",", "").replace("*", "").replace("＊", "")
    try:
        return int(float(s))
    except ValueError:
        return None


def get_year_from_sheet(mlit_df):
    """令和X年Y月 の行から西暦年を取得"""
    for i in range(4, 10):
        cell = str(mlit_df.iloc[i, 0]).strip()
        # 令和表記
        m = re.search(r'令和(\d+)年(\d+)月', cell)
        if m:
            reiwa = int(m.group(1))
            return 2018 + reiwa
        # 西暦表記
        m = re.search(r'(\d{4})年(\d+)月', cell)
        if m:
            return int(m.group(1))
    return None


def build_new_rows(mlit_df, sheet_name, filename):
    """参考第1表データ → 追加用DataFrameを生成"""
    month = parse_month_from_sheet(sheet_name)
    year = get_year_from_sheet(mlit_df)
    if year is None or month is None:
        raise ValueError(f"年月を特定できませんでした (sheet={sheet_name})")
    print(f"   データ年月: {year}年{month}月")

    # 都道府県コード→地図用名称
    def to_map_name(s):
        return re.sub(r'^\d+', '', s.strip())

    rows = []
    for row_idx in range(7, 54):  # 47都道府県
        pref_raw = str(mlit_df.iloc[row_idx, 0]).strip()
        都道府県 = to_map_name(pref_raw)
        for col_offset, (nat_mlit, nat_existing) in enumerate(NATIONALITY_MAP.items()):
            val = parse_value(mlit_df.iloc[row_idx, 2 + col_offset])
            rows.append({
                "延べ宿泊数": val,
                "年月": pd.Timestamp(f"{year}-{month:02d}-01"),
                "施設所在地": pref_raw,
                "都道府県（地図用）": 都道府県,
                "国籍": nat_existing,
                "年": year,
                "月": month,
                "Table Names-1": f"{filename}/参考第1表({month}月)",
                "Table Names": f"参考第1表({month}月)",
                "File Paths": filename,
            })
    return pd.DataFrame(rows), year, month


def save_to_dl_folder(new_df):
    """縦持ちデータを日付付きファイル名でDLフォルダに保存（既存ファイルは上書きしない）"""
    import os
    os.makedirs(DL_FOLDER, exist_ok=True)
    today = datetime.now().strftime("%Y%m%d")
    out_path = os.path.join(DL_FOLDER, f"{today}_宿泊データ縦持ち.xlsx")

    if os.path.exists(out_path):
        print(f"[3b] スキップ: 本日分ファイルが既に存在します → {os.path.basename(out_path)}")
        return out_path

    # 保存列（縦持ち形式の主要列のみ）
    save_cols = ["延べ宿泊数", "年月", "施設所在地", "都道府県（地図用）", "国籍", "年", "月",
                 "Table Names-1", "Table Names", "File Paths"]
    df_save = new_df[save_cols].copy()

    df_save.to_excel(out_path, index=False, sheet_name="元データ")
    print(f"[3b] DLフォルダに保存: {os.path.basename(out_path)}  ({len(df_save)}行)")
    return out_path


def update_excel(new_df, year, month):
    """既存Excelの同月データを削除し、新データを追記"""
    print(f"[4] Excelファイルを更新中: {EXCEL_FILE}")
    wb = load_workbook(EXCEL_FILE)
    ws = wb.active

    # ヘッダー行確認
    header = [ws.cell(1, c).value for c in range(1, 11)]
    col_map = {v: i+1 for i, v in enumerate(header) if v}

    year_col = col_map.get("年")
    month_col = col_map.get("月")

    # 同月の行を削除（後ろから走査）
    deleted = 0
    for row in range(ws.max_row, 1, -1):
        y = ws.cell(row, year_col).value
        m = ws.cell(row, month_col).value
        if y == year and m == month:
            ws.delete_rows(row)
            deleted += 1

    print(f"   削除した既存行: {deleted}行")

    # 新データを追記
    cols = ["延べ宿泊数", "年月", "施設所在地", "都道府県（地図用）", "国籍",
            "年", "月", "Table Names-1", "Table Names", "File Paths"]
    for _, row in new_df.iterrows():
        row_data = []
        for col in cols:
            val = row[col]
            if col == "年月":
                row_data.append(val.to_pydatetime())
            elif col in ("年", "月", "延べ宿泊数"):
                row_data.append(int(val) if pd.notna(val) else None)
            else:
                row_data.append(val)
        ws.append(row_data)

    wb.save(EXCEL_FILE)
    print(f"   追加行: {len(new_df)}行")
    print(f"   保存完了: {EXCEL_FILE}")


def verify(new_df, year, month, n=15):
    """ランダムN箇所でMLITデータと保存済みデータを照合"""
    print(f"[5] 検証中（ランダム{n}箇所）...")
    df = pd.read_excel(EXCEL_FILE, sheet_name="元データ", header=0)
    df_target = df[(df["年"] == year) & (df["月"] == month)]

    existing_lookup = {
        (row["施設所在地"], row["国籍"]): row["延べ宿泊数"]
        for _, row in df_target.iterrows()
    }
    source_lookup = {
        (row["施設所在地"], row["国籍"]): row["延べ宿泊数"]
        for _, row in new_df.iterrows()
    }

    all_keys = list(source_lookup.keys())
    sample = random.sample(all_keys, min(n, len(all_keys)))

    header = f"{'#':<3} {'施設所在地':<12} {'国籍':<14} {'元データ':>10} {'保存値':>10} {'結果'}"
    print(header)
    print("-" * 60)

    ng_count = 0
    for i, key in enumerate(sample, 1):
        pref, nat = key
        src = source_lookup.get(key)
        dst = existing_lookup.get(key)
        ok = src == dst
        mark = "✓" if ok else "✗ NG"
        if not ok:
            ng_count += 1
        print(f"{i:<3} {pref:<12} {nat:<14} {str(src):>10} {str(dst):>10} {mark}")

    print()
    if ng_count == 0:
        print(f"✅ 全{n}箇所一致。データは正確に更新されました。")
    else:
        print(f"❌ {ng_count}箇所で不一致が見つかりました。要確認。")
        sys.exit(1)


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--dry-run", action="store_true", help="ファイルを更新せず確認のみ")
    parser.add_argument("--url", help="ダウンロードURLを直接指定（省略時はMLITから最新を自動取得）")
    args = parser.parse_args()

    print("=" * 60)
    print(f"宿泊旅行統計 参考第１表 月次更新スクリプト")
    print(f"実行日時: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print(f"更新対象ファイル: {EXCEL_FILE}")
    print("=" * 60)

    # 1. ファイルURLを決定
    if args.url:
        url = args.url
        filename = args.url.split("/")[-1]
        print(f"[1] 指定URLを使用: {url}")
    else:
        url, filename = fetch_latest_excel_url()

    # 2. ダウンロード
    xlsx_bytes = download_excel(url)

    # 3. 参考第1表シートを読み込む
    print("[3] 参考第1表を解析中...")
    xl = pd.ExcelFile(io.BytesIO(xlsx_bytes))
    sheet_name = find_ref1_sheet(xl.sheet_names)
    if sheet_name is None:
        raise RuntimeError(f"参考第1表シートが見つかりません。シート一覧: {xl.sheet_names}")
    print(f"   対象シート: {sheet_name}")
    mlit_df = pd.read_excel(io.BytesIO(xlsx_bytes), sheet_name=sheet_name, header=None)

    new_df, year, month = build_new_rows(mlit_df, sheet_name, filename)
    print(f"   生成行数: {len(new_df)}行 ({year}年{month}月 × 47都道府県 × 24国籍)")

    if args.dry_run:
        print("\n[DRY RUN] ファイルの更新はスキップしました。")
        print(f"追加予定データのサンプル:")
        print(new_df.head(5).to_string())
        return

    # 3b. DLフォルダに縦持ちファイルを日付付きで保存
    save_to_dl_folder(new_df)

    # 4. 元データExcelを更新
    update_excel(new_df, year, month)

    # 5. 検証
    verify(new_df, year, month, n=15)

    print()
    print(f"🎉 完了: {year}年{month}月分データを更新しました。")


if __name__ == "__main__":
    main()
