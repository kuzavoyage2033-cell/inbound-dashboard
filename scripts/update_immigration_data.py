"""
出入国管理統計 港別入国外国人の国籍・地域 自動更新スクリプト

e-Statから新しい月報データを確認し、未取得分をダウンロードして
既存のExcelファイルに追加します。
"""

import os
import re
import shutil
import tempfile
import urllib.request
import urllib.parse
import html
from datetime import datetime, date
from pathlib import Path
import openpyxl
import pandas as pd


# ===== 設定 =====
BASE_DIR = str(Path(__file__).resolve().parent.parent)
DATA_DIR = str(Path(BASE_DIR).parent.parent / "00_観光データ格納庫")
EXCEL_OUTPUT = os.path.join(DATA_DIR, "入国データ自動DL", "港別_入国外国人_国籍地域_縦持ち.xlsx")
DOWNLOAD_DIR = os.path.join(BASE_DIR, "archive", "immigration_data")
ESTAT_BASE = "https://www.e-stat.go.jp"
HEADERS = {
    "User-Agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36",
    "Referer": "https://www.e-stat.go.jp/"
}


def fetch_url(url):
    req = urllib.request.Request(url, headers=HEADERS)
    with urllib.request.urlopen(req, timeout=30) as resp:
        return resp.read().decode("utf-8", errors="replace")


def get_existing_months():
    """既存Excelファイルから取得済みの年月を返す"""
    if not os.path.exists(EXCEL_OUTPUT):
        return set()
    df = pd.read_excel(EXCEL_OUTPUT, usecols=["年", "月"])
    months = set(zip(df["年"].astype(int), df["月"].astype(int)))
    return months


def find_stat_inf_id(year, month):
    """
    e-Statのページから指定年月の「港別入国外国人の国籍・地域」のstatInfIdを探す
    """
    url = (
        f"{ESTAT_BASE}/stat-search/files"
        f"?page=1&layout=dataset&cycle=1"
        f"&year={year}0&toukei=00250011"
        f"&tstat=000001012480&tclass1=000001012481"
    )

    # 複数ページを検索
    for page in range(1, 6):
        page_url = url.replace("page=1", f"page={page}")
        try:
            content = fetch_url(page_url)
        except Exception as e:
            print(f"  ページ取得エラー (page={page}): {e}")
            break

        # 対象月の表番号パターン: YY-MM-01-2 (港別入国外国人の国籍・地域)
        yy = str(year)[2:]
        mm = str(month).zfill(2)
        pattern = f"{yy}-{mm}-01-2"

        if pattern not in content:
            continue

        # statInfIdを抽出 (file-download?statInfId=XXXXXXXXX の形式)
        # パターンの前後のコンテキストを探す
        idx = content.find(pattern)
        # 前後2000文字を検索してstatInfIdを探す
        context = content[max(0, idx-2000):idx+2000]

        # 閲覧用(fileKind=4)のstatInfIdを探す
        matches = re.findall(r'statInfId=(\d{12})&fileKind=4', context)
        if matches:
            return matches[0]

        # fileKind=4が見つからない場合はfileKind=0も試す
        matches = re.findall(r'statInfId=(\d{12})&fileKind=0', context)
        if matches:
            return matches[0]

    return None


def download_excel(stat_inf_id, year, month):
    """ExcelファイルをダウンロードしてDOWNLOAD_DIRに保存"""
    url = f"{ESTAT_BASE}/stat-search/file-download?statInfId={stat_inf_id}&fileKind=4"
    fname = f"{year}_{str(month).zfill(2)}.xlsx"
    outpath = os.path.join(DOWNLOAD_DIR, fname)

    req = urllib.request.Request(url, headers=HEADERS)
    with urllib.request.urlopen(req, timeout=30) as resp:
        data = resp.read()

    with open(outpath, "wb") as f:
        f.write(data)

    return outpath


def verify_file(path, year, month):
    """ダウンロードしたファイルが正しい表かチェック"""
    try:
        wb = openpyxl.load_workbook(path, data_only=True)
        ws = wb.active
        title = ws.cell(1, 1).value or ""
        return "港別" in str(title) and "入国外国人" in str(title) and "国籍" in str(title)
    except Exception:
        return False


def parse_excel_to_records(path, year, month):
    """ExcelファイルをパースしてDataFrame用レコードリストを返す"""
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active

    # 4行目: 港名
    ports = []
    for c in range(1, ws.max_column + 1):
        val = ws.cell(4, c).value
        if val is not None:
            clean_val = str(val).replace("\n", "").strip()
            ports.append((c, clean_val))

    records = []
    for r in range(5, ws.max_row + 1):
        nationality = ws.cell(r, 1).value
        if nationality is None:
            continue
        nationality = str(nationality).strip()
        if not nationality:
            continue

        for col_idx, port_name in ports:
            if col_idx == 1:
                continue
            val = ws.cell(r, col_idx).value
            if val is None:
                val = 0
            records.append({
                "年": year,
                "月": month,
                "国籍・地域": nationality,
                "港": port_name,
                "入国外国人数": val
            })

    return records


def append_to_excel(new_records):
    """既存Excelファイルに新しいレコードを追加"""
    new_df = pd.DataFrame(new_records)

    if os.path.exists(EXCEL_OUTPUT):
        existing_df = pd.read_excel(EXCEL_OUTPUT)
        combined_df = pd.concat([existing_df, new_df], ignore_index=True)
        combined_df = combined_df.sort_values(["年", "月", "国籍・地域", "港"]).reset_index(drop=True)
    else:
        combined_df = new_df.sort_values(["年", "月", "国籍・地域", "港"]).reset_index(drop=True)

    # iCloud等ネットワークドライブへの直接書き込みを避けるため一時ファイルに書いてから移動
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
        tmp_path = tmp.name

    try:
        with pd.ExcelWriter(tmp_path, engine="xlsxwriter") as writer:
            combined_df.to_excel(writer, index=False, sheet_name="港別入国外国人_縦持ち")
            worksheet = writer.sheets["港別入国外国人_縦持ち"]
            worksheet.set_column("A:A", 6)
            worksheet.set_column("B:B", 6)
            worksheet.set_column("C:C", 30)
            worksheet.set_column("D:D", 20)
            worksheet.set_column("E:E", 15)
        shutil.move(tmp_path, EXCEL_OUTPUT)
    except Exception:
        if os.path.exists(tmp_path):
            os.remove(tmp_path)
        raise

    return len(combined_df)


def main():
    print(f"=== 出入国管理統計 自動更新 {datetime.now().strftime('%Y-%m-%d %H:%M')} ===")

    os.makedirs(DOWNLOAD_DIR, exist_ok=True)

    # 取得済み年月を確認
    existing = get_existing_months()
    print(f"取得済み月数: {len(existing)}")
    if existing:
        last = max(existing)
        print(f"最新取得済み: {last[0]}年{last[1]}月")

    # 確認対象: 2024年1月から現在の2ヶ月前まで
    today = date.today()
    # 月次データは約2ヶ月後に公開されるため、2ヶ月前まで確認
    check_until_year = today.year
    check_until_month = today.month - 2
    if check_until_month <= 0:
        check_until_month += 12
        check_until_year -= 1

    new_months_added = []

    # 2024年1月から確認
    year, month = 2024, 1
    while (year, month) <= (check_until_year, check_until_month):
        if (year, month) not in existing:
            print(f"\n新しいデータを確認中: {year}年{month}月...")

            # ローカルに既にファイルがあればダウンロードをスキップ
            local_path = os.path.join(DOWNLOAD_DIR, f"{year}_{str(month).zfill(2)}.xlsx")
            if os.path.exists(local_path):
                print(f"  ローカルファイルを使用: {os.path.basename(local_path)}")
                try:
                    if verify_file(local_path, year, month):
                        records = parse_excel_to_records(local_path, year, month)
                        total = append_to_excel(records)
                        print(f"  追加完了: {len(records)}行追加 (合計{total}行)")
                        new_months_added.append(f"{year}年{month}月")
                    else:
                        print(f"  ファイル検証失敗: 正しい表ではない可能性があります")
                except Exception as e:
                    print(f"  処理エラー: {e}")
            else:
                stat_id = find_stat_inf_id(year, month)
                if stat_id:
                    print(f"  statInfId発見: {stat_id}")
                    try:
                        dl_path = download_excel(stat_id, year, month)
                        if verify_file(dl_path, year, month):
                            records = parse_excel_to_records(dl_path, year, month)
                            total = append_to_excel(records)
                            print(f"  追加完了: {len(records)}行追加 (合計{total}行)")
                            new_months_added.append(f"{year}年{month}月")
                        else:
                            print(f"  ファイル検証失敗: 正しい表ではない可能性があります")
                            os.remove(dl_path)
                    except Exception as e:
                        print(f"  ダウンロード/処理エラー: {e}")
                else:
                    print(f"  データ未公開またはID未発見")

        # 次の月へ
        month += 1
        if month > 12:
            month = 1
            year += 1

    if new_months_added:
        print(f"\n=== 更新完了: {', '.join(new_months_added)} を追加しました ===")
    else:
        print(f"\n=== 更新なし: 新しいデータはありません ===")

    print(f"出力ファイル: {EXCEL_OUTPUT}")
    return new_months_added


if __name__ == "__main__":
    main()
