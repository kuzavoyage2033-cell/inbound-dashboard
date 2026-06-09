"""
出入国管理統計 港別出国外国人の国籍・地域 スクリプト

e-Statから2024年1月〜2026年2月（または最新公開データ）をダウンロードして
縦持ち形式のExcelファイルに変換して保存します。

保存先: 01_観光データ分析/出国データ自動DL/yyyymmdd_出国データ縦持ち.xlsx
"""

import os
import re
import shutil
import tempfile
import urllib.request
from datetime import datetime, date
from pathlib import Path
import openpyxl
import pandas as pd


# ===== 設定 =====
BASE_DIR = str(Path(__file__).resolve().parent.parent)
DATA_DIR = str(Path(BASE_DIR).parent.parent / "00_観光データ格納庫")
OUTPUT_DIR = os.path.join(DATA_DIR, "出国データ自動DL")
DOWNLOAD_DIR = os.path.join(BASE_DIR, "archive", "emigration_data_raw")
ESTAT_BASE = "https://www.e-stat.go.jp"
HEADERS = {
    "User-Agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36",
    "Referer": "https://www.e-stat.go.jp/"
}

# データ取得範囲
START_YEAR = 2024
START_MONTH = 1


def get_output_path():
    """今日の日付を使った出力ファイルパスを返す"""
    today = datetime.now().strftime("%Y%m%d")
    return os.path.join(OUTPUT_DIR, f"{today}_出国データ縦持ち.xlsx")


def fetch_url(url):
    req = urllib.request.Request(url, headers=HEADERS)
    with urllib.request.urlopen(req, timeout=30) as resp:
        return resp.read().decode("utf-8", errors="replace")


def find_stat_inf_id(year, month):
    """
    e-Statのページから指定年月の「港別出国外国人の国籍・地域」のstatInfIdを探す。
    表番号パターン: {yy}-{mm}-02-2（出国外国人、入国は01-2）
    """
    url = (
        f"{ESTAT_BASE}/stat-search/files"
        f"?page=1&layout=dataset&cycle=1"
        f"&year={year}0&toukei=00250011"
        f"&tstat=000001012480&tclass1=000001012481"
    )

    yy = str(year)[2:]
    mm = str(month).zfill(2)
    # 出国データの表番号: YY-MM-01-3（港別出国外国人の国籍・地域）
    pattern = f"{yy}-{mm}-01-3"

    for page in range(1, 9):
        page_url = url.replace("page=1", f"page={page}")
        try:
            content = fetch_url(page_url)
        except Exception as e:
            print(f"  ページ取得エラー (page={page}): {e}")
            break

        if pattern not in content:
            continue

        idx = content.find(pattern)
        context = content[max(0, idx - 2000):idx + 2000]

        matches = re.findall(r'statInfId=(\d{12})&fileKind=4', context)
        if matches:
            return matches[0]

        # fileKind=4が見つからない場合はfileKind=0を試す
        matches = re.findall(r'statInfId=(\d{12})&fileKind=0', context)
        if matches:
            return matches[0]

    return None


def download_excel(stat_inf_id, year, month):
    """ExcelファイルをダウンロードしてDOWNLOAD_DIRに保存"""
    url = f"{ESTAT_BASE}/stat-search/file-download?statInfId={stat_inf_id}&fileKind=4"
    fname = f"{year}_{str(month).zfill(2)}_emigration.xlsx"
    outpath = os.path.join(DOWNLOAD_DIR, fname)

    req = urllib.request.Request(url, headers=HEADERS)
    with urllib.request.urlopen(req, timeout=30) as resp:
        data = resp.read()

    with open(outpath, "wb") as f:
        f.write(data)

    return outpath


def verify_file(path):
    """ダウンロードしたファイルが「港別出国外国人の国籍・地域」の表かチェック"""
    try:
        wb = openpyxl.load_workbook(path, data_only=True)
        ws = wb.active
        title = ws.cell(1, 1).value or ""
        return "港別" in str(title) and "出国外国人" in str(title) and "国籍" in str(title)
    except Exception:
        return False


def parse_excel_to_records(path, year, month):
    """ExcelファイルをパースしてDataFrame用レコードリストを返す"""
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active

    # 4行目: 港名ヘッダー
    ports = []
    for c in range(1, ws.max_column + 1):
        val = ws.cell(4, c).value
        if val is not None:
            clean_val = str(val).replace("\n", "").strip()
            ports.append((c, clean_val))

    records = []
    for r in range(5, ws.max_row + 1):
        nationality = ws.cell(r, 1).value
        if nationality is None:
            continue
        nationality = str(nationality).strip()
        if not nationality:
            continue

        for col_idx, port_name in ports:
            if col_idx == 1:
                continue
            val = ws.cell(r, col_idx).value
            if val is None:
                val = 0
            records.append({
                "年": year,
                "月": month,
                "国籍・地域": nationality,
                "港": port_name,
                "出国外国人数": val
            })

    return records


def save_to_excel(all_records, output_path):
    """全レコードを縦持ちExcelとして保存"""
    df = pd.DataFrame(all_records)
    df = df.sort_values(["年", "月", "国籍・地域", "港"]).reset_index(drop=True)

    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
        tmp_path = tmp.name

    try:
        with pd.ExcelWriter(tmp_path, engine="xlsxwriter") as writer:
            df.to_excel(writer, index=False, sheet_name="港別出国外国人_縦持ち")
            ws = writer.sheets["港別出国外国人_縦持ち"]
            ws.set_column("A:A", 6)
            ws.set_column("B:B", 6)
            ws.set_column("C:C", 30)
            ws.set_column("D:D", 20)
            ws.set_column("E:E", 15)
        shutil.move(tmp_path, output_path)
    except Exception:
        if os.path.exists(tmp_path):
            os.remove(tmp_path)
        raise

    return len(df)


def calc_check_until():
    """取得上限年月を返す（月次データは約2ヶ月後に公開）"""
    today = date.today()
    pub_year = today.year
    pub_month = today.month - 2
    if pub_month <= 0:
        pub_month += 12
        pub_year -= 1
    return pub_year, pub_month


def main():
    print(f"=== 出入国管理統計 出国データ更新 {datetime.now().strftime('%Y-%m-%d %H:%M')} ===")

    output_path = get_output_path()

    # 同日ファイルが既に存在する場合はスキップ（履歴保持）
    if os.path.exists(output_path):
        print(f"本日のファイルは既に存在します: {os.path.basename(output_path)}")
        print("既存ファイルは上書きしません。処理を終了します。")
        return []

    os.makedirs(OUTPUT_DIR, exist_ok=True)
    os.makedirs(DOWNLOAD_DIR, exist_ok=True)

    check_until_year, check_until_month = calc_check_until()
    print(f"取得範囲: {START_YEAR}年{START_MONTH}月 〜 {check_until_year}年{check_until_month}月")

    all_records = []
    months_collected = []
    months_failed = []

    year, month = START_YEAR, START_MONTH
    while (year, month) <= (check_until_year, check_until_month):
        print(f"\n処理中: {year}年{month}月...")

        local_path = os.path.join(DOWNLOAD_DIR, f"{year}_{str(month).zfill(2)}_emigration.xlsx")

        if os.path.exists(local_path):
            print(f"  ローカルキャッシュを使用: {os.path.basename(local_path)}")
            if verify_file(local_path):
                records = parse_excel_to_records(local_path, year, month)
                all_records.extend(records)
                months_collected.append(f"{year}年{month}月")
                print(f"  読込完了: {len(records)}行")
            else:
                print(f"  ファイル検証失敗: 出国データではない可能性があります")
                months_failed.append(f"{year}年{month}月")
        else:
            stat_id = find_stat_inf_id(year, month)
            if stat_id:
                print(f"  statInfId発見: {stat_id}")
                try:
                    dl_path = download_excel(stat_id, year, month)
                    if verify_file(dl_path):
                        records = parse_excel_to_records(dl_path, year, month)
                        all_records.extend(records)
                        months_collected.append(f"{year}年{month}月")
                        print(f"  取得完了: {len(records)}行")
                    else:
                        print(f"  ファイル検証失敗: 正しい表ではない可能性があります")
                        os.remove(dl_path)
                        months_failed.append(f"{year}年{month}月")
                except Exception as e:
                    print(f"  ダウンロード/処理エラー: {e}")
                    months_failed.append(f"{year}年{month}月")
            else:
                print(f"  データ未公開またはID未発見")
                months_failed.append(f"{year}年{month}月")

        month += 1
        if month > 12:
            month = 1
            year += 1

    if all_records:
        print(f"\n縦持ちExcelを保存中... ({len(all_records)}行)")
        total = save_to_excel(all_records, output_path)
        print(f"\n=== 完了: {len(months_collected)}ヶ月分 ({total}行) を保存しました ===")
        print(f"出力ファイル: {output_path}")
        if months_failed:
            print(f"取得失敗: {', '.join(months_failed)}")
    else:
        print(f"\n=== データが取得できませんでした ===")
        if months_failed:
            print(f"失敗月: {', '.join(months_failed)}")

    return months_collected


if __name__ == "__main__":
    main()
